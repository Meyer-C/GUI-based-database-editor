import dearpygui.dearpygui as dpg
import openpyxl
import xlsxwriter
import sys
import os

from GUI_compatible_searches import *
from Compound_From_XLS import get_compounds as gc
from openpyxl import load_workbook
from PubChem_Search import search_pubchem_input as spc
from hmdb_xml_data_extactor import *

class windowCount:
    manual_count = 0
    search_count = 0
    pubchem_count = 0
    search_popup_count = 0
    pubchem_popup = 0
    invalid_filepath_popup_count = 0
    output_tab_count = 0
    hmdb_search_count = 0
    hmdb_popup_count = 0
    xlsx_out_count = 0
    beautiful_output_count = 0
    file_window_count = 0
    personal_path_popup = 0
    close_popup_count = 0
    help_count = 0

class namingCount:
    count = 1

class NewChemical:
    info = []


# ---------------------------------------------------------------------------------------------------------------------
# All fxns and classes of search tabs
def search_starter():
    windowCount.output_tab_count += 1
    output_tab_tag = str(f'Outputs {windowCount.output_tab_count}')
    with dpg.tab(label='Outputs', parent='tabs', tag=output_tab_tag):
        search_type = dpg.get_value(f'type_{windowCount.search_count}')
        search_term = dpg.get_value(f'term_{windowCount.search_count}')
        quick_search_str = GUI_quick_search(search_type, search_term, get_filepath(), None, 'str')
        quick_search_compounds = GUI_quick_search(search_type, search_term, get_filepath(), None, 'cmp')
        if quick_search_str[0] != 0:
            compound_tags = []
            for x, compound_str in enumerate(quick_search_str):
                compound_tag = str(f'{search_term}_{x}')
                compound_tags.append(compound_tag)
                dpg.add_checkbox(label=compound_str, tag=compound_tag, parent=output_tab_tag)
            more_info_tag = str(f'{output_tab_tag}_more_button')
            export_tag = str(f'{output_tab_tag}_export_button')
            dpg.add_button(label='More Information(pending)', tag=more_info_tag)
            dpg.add_button(label='Export Data to .xlsx', tag=export_tag, callback=xlsx_out_tab,
                           user_data=[compound_tags, quick_search_compounds])
            dpg.add_button(label='Close', parent=output_tab_tag, callback=close, user_data=output_tab_tag)

def advanced_search_starter(sender, app_data, user_data):
    advanced_return_nums = []
    for x in range(len(user_data)):
        if dpg.get_value(user_data[x]) == True:
            advanced_return_nums.append(x)
    windowCount.output_tab_count += 1
    output_tab_tag = str(f'Outputs {windowCount.output_tab_count}')
    with dpg.tab(label='Outputs', tag=output_tab_tag, parent='tabs'):
        search_type = dpg.get_value(f'type_{windowCount.search_count}')
        search_term = dpg.get_value(f'term_{windowCount.search_count}')
        quick_search_str = GUI_quick_search(search_type, search_term, get_filepath(), advanced_return_nums, 'str')
        quick_search_compounds = GUI_quick_search(search_type, search_term, get_filepath(), advanced_return_nums, 'cmp')
        if quick_search_str[0] != 0:
            compound_tags = []
            compounds = []
            for x, compound_str in enumerate(quick_search_str):
                compounds.append(quick_search_compounds[x])
                compound_tag = str(f'{search_term}_{x}')
                compound_tags.append(compound_tag)
                dpg.add_checkbox(label=compound_str, tag=compound_tag, parent=output_tab_tag)
            more_info_tag = str(f'{output_tab_tag}_more_button')
            export_tag = str(f'{output_tab_tag}_export_button')
            dpg.add_button(label='More Information', tag=more_info_tag, callback=beautiful_output_tab,
                           user_data=[compound_tags, compounds])
            dpg.add_button(label='Export Data to .xlsx', tag=export_tag, callback=xlsx_out_tab,
                           user_data=[compound_tags, compounds])
            dpg.add_button(label='Close', parent=output_tab_tag, callback=close, user_data=output_tab_tag)
            close(sender, app_data, str(f'Advanced Search {windowCount.search_popup_count}'))

class searchWindow():

    def search_window(self):
        if gc(get_filepath()) == 'Invalid filepath':
            invalid_filepath_popup()
        else:
            search_types = ['Name', 'Mass', 'InChiKey']
            windowCount.search_count += 1
            tab_tag = str(f'Search {windowCount.search_count}')
            with dpg.tab(label=tab_tag, tag=tab_tag, parent='tabs'):
                search_type_call = str(f'type_{windowCount.search_count}')
                search_term_call = str(f'term_{windowCount.search_count}')
                dpg.add_radio_button(tag=search_type_call, items=search_types, horizontal=True,
                                     default_value=search_types[0], parent=tab_tag)
                dpg.add_input_text(tag=search_term_call, label='Search Here', parent=tab_tag)
                dpg.add_button(label='Advanced Search Options', callback=searchWindowPopup.search_window_popup)
                dpg.add_button(label='Search!', callback=search_starter, parent=tab_tag)
                dpg.add_button(label='X', pos=(dpg.get_viewport_width() - 40, 50), parent=tab_tag, callback=close,
                               user_data=tab_tag)

class searchWindowPopup():

    def search_window_popup(self):
        windowCount.search_popup_count += 1
        window_name = str(f'Advanced Search {windowCount.search_popup_count}')
        with dpg.window(tag=window_name, width=400, height=300):
            compound_list = gc(get_filepath())
            names = compound_list[0]
            name_tags = []
            for x in range(len(names)):
                name = names[x]
                if name is not None:
                    name_tag = str(f'{x}. {names[x]} {windowCount.search_popup_count}')
                    name_tags.append(name_tag)
                    dpg.add_checkbox(label=name, tag=name_tag, parent=window_name)
            dpg.add_button(label='Advanced Search!', callback=advanced_search_starter, user_data=name_tags,
                           parent=window_name)

def xlsx_out_tab(sender, app_data, user_data):
    windowCount.xlsx_out_count += 1
    window_name = str(f'Write compound to .xlsx {windowCount.xlsx_out_count}')
    compound_tags = user_data[0]
    compounds = user_data[1]
    compounds_to_export = []
    with dpg.window(tag=window_name, width=600, height=300):
        for x, compound_tag in enumerate(compound_tags):
            if dpg.get_value(compound_tag) == True:
                compounds_to_export.append(compounds[x])
        if len(compounds_to_export) != 0:
            path_info = get_filepath(False)
            path_tags = {}
            dpg.add_text('Select one of your existing files or create a new one.')
            for path_name, path in path_info.items():
                path_tag = str(f'{path_name}_{windowCount.xlsx_out_count}')
                path_tags[path_tag] = path
                dpg.add_checkbox(label=path_name, tag=path_tag, parent=window_name)
            filepath_input_tag = str(f'filepath_input_{windowCount.xlsx_out_count}')
            name_tag = str(f'name_input_{windowCount.xlsx_out_count}')
            dpg.add_text('New database name:', parent=window_name)
            dpg.add_input_text(tag=name_tag, parent=window_name)
            dpg.add_text('New database filepath and file name:', parent=window_name)
            dpg.add_input_text(parent=window_name, tag=filepath_input_tag)
            #### working to make this callback so that we can have existing filepaths/databases selected
            dpg.add_button(label='Write data to the excel file!', parent=window_name, callback=write_new_xlsx,
                           user_data=(compounds_to_export, filepath_input_tag, path_tags, window_name, name_tag))
        else:
            dpg.add_text('No Compounds Selected', parent=window_name)
            dpg.add_button(label='Close', parent=window_name, callback=close, user_data=window_name)

def write_new_xlsx(sender, app_data, user_data):
    compounds = user_data[0]
    naming_info = gc(get_filepath())[0]
    out_filepath = dpg.get_value(user_data[1])
    existing_out_filepaths = user_data[2]
    name = dpg.get_value(user_data[4])
    new_out_filepaths = []
    final_path = ''

    # TODO work on using existing files and tags
    for path_tag, path in existing_out_filepaths.items():
        if dpg.get_value(path_tag) == True:
            new_out_filepaths.append(path)

    if len(new_out_filepaths) == 0:
        if out_filepath == '' and name == '':
            close_popup(message='NO FILEPATH AND NAME ENTERED\nPlease enter a new output filepath and name or select an'
                                ' existing filepath.')
        elif out_filepath == '' and name != '':
            close_popup(message='NO FILEPATH ENTERED\nPlease enter a new output filepath.')
        elif out_filepath != '' and name == '':
            close_popup(message='NO DATABASE NAME ENTERED\nPlease enter a new database name')
        else:
            if out_filepath.endswith('.xlsx'):
                add_path_no_button(name, out_filepath)
            else:
                final_path += str(f'{out_filepath}.xlsx')
                add_path_no_button(name, final_path)
    else:
        final_path += new_out_filepaths[0]

    try:
        wb = openpyxl.load_workbook(final_path)
        sheet = wb.active
        for x, info in enumerate(naming_info):
            if sheet.cell(1, x + 1).value != naming_info[x]:
                sheet.cell(1, x + 1).value = naming_info[x]
        wb.save(final_path)
        new_row = sheet.max_row + 1
        for y, compound_info in enumerate(compounds):
            for z, this_data in enumerate(compound_info):
                sheet.cell(new_row + y, z + 1).value = this_data
        wb.save(final_path)

    except:
        workbook = xlsxwriter.Workbook(final_path)
        workbook.add_worksheet()
        workbook.close()

        wb = openpyxl.load_workbook(final_path)
        sheet = wb.active
        for x, info in enumerate(naming_info):
            if sheet.cell(1, x + 1).value != naming_info[x]:
                sheet.cell(1, x + 1).value = naming_info[x]
        wb.save(final_path)
        new_row = sheet.max_row + 1
        for y, compound_info in enumerate(compounds):
            for z, this_data in enumerate(compound_info):
                sheet.cell(new_row + y, z + 1).value = this_data
        wb.save(final_path)

    finally:
        close_popup(message='Chemicals have been added to your database!', tab_tag=user_data[3])

def beautiful_output_tab(sender, app_data, user_data):
    compound_tags = user_data[0]
    compounds = user_data[1]
    compounds_to_display = []
    windowCount.beautiful_output_count += 1
    naming_info = gc(get_filepath())[0]
    for x, compound_tag in enumerate(compound_tags):
        if dpg.get_value(compound_tag) == True:
            compounds_to_display.append(compounds[x])
    for this_compound in compounds_to_display:
        tab_tag = str(f'{windowCount.beautiful_output_count} {this_compound[1]}')
        dpg.add_tab(label=tab_tag, tag=tab_tag, parent='tabs')
        compound_info_str = ''
        checkbox_tag = str(f'{tab_tag}_check')
        for y, name in enumerate(naming_info):
            if this_compound[y] is not None and naming_info is not None:
                compound_info_str += str(f'{str(naming_info[y]):40} {str(this_compound[y])}\n')
        dpg.add_checkbox(label=compound_info_str, tag=checkbox_tag, parent=tab_tag)
        dpg.add_button(label='Write data to the excel file!', parent=tab_tag, callback=xlsx_out_tab,
                       user_data=[[checkbox_tag], [this_compound]])
        dpg.add_button(label='Close', parent=tab_tag, callback=close, user_data=tab_tag)


# --------------------------------------------------------------------------------------------------------------------
# all classes and fxns of manual chemical addition

def update_input_text(sender, value, user_data):
        compound_list = gc(get_filepath())
        if compound_list != 'Invalid filepath':
            names = compound_list[0]
            if sender == 'f':
                namingCount.count += 1
                if names[namingCount.count] is not None:
                    dpg.set_value(user_data, names[namingCount.count])
                else:
                    dpg.add_button(label='Add Chemical!', parent='manual_add')
            elif sender == 'sv':
                print(dpg.get_value(namingCount.count))
            else:
                namingCount.count -= 1
                dpg.set_value(user_data, names[namingCount.count])

        else:
            dpg.set_value(user_data, 'Invalid Filepath')

class manualTabAdd:

    def manual_window(self):
        if gc(get_filepath()) == 'Invalid filepath':
            invalid_filepath_popup()
        else:
            filepath = get_filepath()
            windowCount.manual_count += 1
            tab_tag = str(f'Manual Chemical Add {str(windowCount.manual_count)}')
            with dpg.tab(tag=tab_tag, label=tab_tag, parent='tabs'):
                compounds = gc(filepath)
                if compounds == 'Invalid Filepath':
                    dpg.add_text('Invalid Filepath')
                else:
                    names = compounds[0]
                    name_tags = []
                    for x in range(1,len(names)):
                        if names[x] is not None and names[x] != 'STANDARD PROPERTIES <<--' and names[x] != \
                                'PARENT COMPOUNT PROPERTIES -->>':
                            name_tag = f'{windowCount.manual_count}-{x}. {names[x]}'
                            name_tags.append(name_tag)
                            dpg.add_text(name_tag, parent=tab_tag)
                            dpg.add_input_text(tag=name_tag, parent=tab_tag)
                    dpg.add_button(label='Add Chemical!', tag=f'{windowCount.manual_count}_add_chemical',
                                   callback=add_to_database, user_data=[name_tags, tab_tag], parent=tab_tag)
            dpg.add_button(tag=str(f'close_{tab_tag}'), label='X', pos=(dpg.get_viewport_width() - 50, 50),
                           parent=tab_tag, callback=close, user_data=tab_tag)

def add_to_database(sender, app_data, user_data):
    name_tags = user_data[0]
    wb = load_workbook(get_filepath())
    sheet = wb.active
    max_row = sheet.max_row
    prev_accession = sheet.cell(max_row, 1).value
    zeros = 6 - len(str(int(prev_accession[5::]) + 1))
    sheet.cell(max_row + 1, 1).value = str(prev_accession[0:5]) + str('0' * zeros) + str(int(prev_accession[5::]) + 1)

    for x in range(len(name_tags)):
        sheet.cell(max_row + 1, x + 2).value = dpg.get_value(name_tags[x])
    wb.save(get_filepath())

    close_popup(tab_tag=user_data[1], message='Chemical has been added manually!')

# ---------------------------------------------------------------------------------------------------------------------
# All classes and fxns of PubChem addition

class pubChemAdd:

    def pubchem_window(self):
        if gc(get_filepath()) == 'Invalid filepath':
            invalid_filepath_popup()
        else:
            windowCount.pubchem_count += 1
            tab_tag = str(f'PubChem Add {windowCount.pubchem_count}')
            search_types = ['Name', 'CID', 'Smiles', 'InChiKey', 'InChi', 'Formula']
            search_type_call = str(f'PubChem_type_{windowCount.pubchem_count}')
            with dpg.tab(tag=tab_tag, label=tab_tag, parent='tabs'):
                dpg.add_radio_button(tag=search_type_call, items=search_types, horizontal=True,
                                     default_value=search_types[0], parent=tab_tag)
                this_search = str(f'{tab_tag} Search')
                dpg.add_input_text(tag=this_search, parent=tab_tag)
                dpg.add_button(tag=str(f'{this_search}_button'),label='Search!', callback=search_pubchem,
                               user_data=[search_type_call, this_search])
            dpg.add_button(label='X', pos=(dpg.get_viewport_width() - 40, 50), parent=tab_tag, callback=close,
                           user_data=tab_tag)

# ASK WHAT KIND OF DATA SHOULD BE RETURNED FROM PUBCHEM?
def search_pubchem(sender, app_data, user_data):
    search_term = dpg.get_value(user_data[1])
    search_type = dpg.get_value(user_data[0]).lower()
    pubchem_popup(spc(search_term, search_type), search_term)


# Maybe do this as a popup to select the compound, then open a tab like the manual add
def pubchem_popup(compounds, search_term):
    windowCount.pubchem_popup += 1
    tab_tag = str(f'{windowCount.pubchem_popup}. Pubchem Search Results: {search_term}')
    with dpg.window(tag=tab_tag, label=tab_tag, width=600, height=300):
        dpg.add_text(label='Select the compounds you want to add!', parent=tab_tag)
        compound_strings = []
        compound_tags = []
        for compound in compounds:
            compound_strings.append(str(f'{"Name":20}{compound.iupac_name}\n{"Formula":20}{compound.molecular_formula}'
                                        f'\n{"InChiKey":20}{compound.inchikey}'))
        for compound_string in compound_strings:
            compound_tag = str(f'{windowCount.pubchem_popup} {compound_string}')
            compound_tags.append(compound_tag)
            dpg.add_checkbox(tag=compound_tag, label=compound_string, parent=tab_tag)
        dpg.add_button(label='Add to Database, Continue to Next Step', parent=tab_tag, callback=pubchem_popup_add_info,
                       user_data=[compound_tags, compounds, tab_tag])

def pubchem_popup_add_info(sender, app_data, user_data):
    dpg.delete_item(sender)
    compounds_to_add = []
    compound_tags = user_data[0]
    compounds = user_data[1]
    tab_tag = user_data[2]
    for x in range(len(compound_tags)):
        if dpg.get_value(compound_tags[x]) == True:
            compounds_to_add.append(compounds[x])
            dpg.delete_item(compound_tags[x])
        else:
            dpg.delete_item(compound_tags[x])

    compounds = gc(get_filepath())
    names = compounds[0]
    name_tags = []
    label_tags = []
    for x in range(1, 13):
        if names[x] is not None and names[x] != 'STANDARD PROPERTIES <<--' and names[x] != \
                'PARENT COMPOUNT PROPERTIES -->>':
            name_tag = str(f'{windowCount.pubchem_popup}. {names[x]}')
            labels = str(f'{x}. {names[x]}')
            label_tag = str(f'{windowCount}_{labels}')
            label_tags.append(label_tag)
            dpg.add_text(labels,tag=label_tag, parent=tab_tag)
            dpg.add_input_text(tag=name_tag, parent=tab_tag)
            name_tags.append(name_tag)
    button_tag = str(f'{windowCount.pubchem_popup}_add_pubchem_chemical')
    dpg.add_button(label='Add Chemical to Database!', tag=button_tag,
                   parent=tab_tag, callback=add_with_pubchem,
                   user_data=[compounds_to_add, name_tags, label_tags, button_tag, tab_tag])

def add_with_pubchem(sender, app_data, user_data):
    compounds_to_add = user_data[0]
    chosen_compound = compounds_to_add[0]
    name_tags = user_data[1]
    label_tags = user_data[2]
    button_tag = user_data[3]
    tab_tag = user_data[4]

    wb = load_workbook(get_filepath())
    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    new_row = max_row + 1

    # creating a new accession
    prev_accession = sheet.cell(max_row, 1).value
    zeros = 6 - len(str(int(prev_accession[5::]) + 1))
    sheet.cell(new_row, 1).value = str(prev_accession[0:5]) + str('0' * zeros) + str(int(prev_accession[5::]) + 1)

    for x in range(2, 14):
        if sheet.cell(1, x).value != None:
            sheet.cell(new_row, x).value = dpg.get_value(name_tags[x-2])

    # just calling all of the different chemical properties from PubChem :\
    sheet.cell(new_row, 14).value = int(chosen_compound.cid)
    sheet.cell(new_row, 15).value = chosen_compound.molecular_formula
    sheet.cell(new_row, 16).value = float(chosen_compound.molecular_weight)
    sheet.cell(new_row, 20).value = int(chosen_compound.cid)
    sheet.cell(new_row, 21).value = chosen_compound.molecular_formula
    sheet.cell(new_row, 22).value = float(chosen_compound.molecular_weight)
    sheet.cell(new_row, 23).value = chosen_compound.canonical_smiles
    sheet.cell(new_row, 24).value = chosen_compound.isomeric_smiles
    sheet.cell(new_row, 25).value = chosen_compound.inchi
    sheet.cell(new_row, 26).value = chosen_compound.inchikey
    sheet.cell(new_row, 27).value = chosen_compound.iupac_name
    # FIXME the excel sheet has a repetitive mass but is listed as iupac name
    sheet.cell(new_row, 29).value = float(chosen_compound.monoisotopic_mass)
    sheet.cell(new_row, 30).value = int(chosen_compound.tpsa)
    sheet.cell(new_row, 31).value = int(chosen_compound.complexity)
    sheet.cell(new_row, 32).value = int(chosen_compound.charge)
    sheet.cell(new_row, 33).value = int(chosen_compound.h_bond_donor_count)
    sheet.cell(new_row, 34).value = int(chosen_compound.h_bond_acceptor_count)
    sheet.cell(new_row, 35).value = int(chosen_compound.rotatable_bond_count)
    sheet.cell(new_row, 36).value = int(chosen_compound.heavy_atom_count)
    sheet.cell(new_row, 37).value = int(chosen_compound.isotope_atom_count)
    sheet.cell(new_row, 38).value = int(chosen_compound.atom_stereo_count)
    sheet.cell(new_row, 39).value = int(chosen_compound.defined_atom_stereo_count)
    sheet.cell(new_row, 40).value = int(chosen_compound.undefined_atom_stereo_count)
    sheet.cell(new_row, 41).value = int(chosen_compound.bond_stereo_count)
    sheet.cell(new_row, 42).value = int(chosen_compound.defined_bond_stereo_count)
    sheet.cell(new_row, 43).value = int(chosen_compound.undefined_bond_stereo_count)
    sheet.cell(new_row, 44).value = int(chosen_compound.covalent_unit_count)
    sheet.cell(new_row, 46).value = chosen_compound.fingerprint
    sheet.cell(new_row, 47).value = float(chosen_compound.xlogp)

    wb.save(get_filepath())

    close_popup(tab_tag=tab_tag, message='Chemical has been added from PubChem!')

class hmdbAdd:
    def hmdb_window(self):
        if get_filepath() == 'Invalid Filepath':
            invalid_filepath_popup()
        else:
            windowCount.hmdb_search_count += 1
            tab_tag = str(f'HMDB Add {windowCount.hmdb_search_count}')
            with dpg.tab(label='HMDB Add', tag=tab_tag, parent='tabs'):
                dpg.add_text('Search the Human Metabolomics Data Base!\nSearch only with the HMDB ID:', parent=tab_tag)
                search_tag = str(f'Search {tab_tag}')
                dpg.add_input_text(tag=search_tag, parent=tab_tag)
                dpg.add_button(label='Search!', parent=tab_tag, callback=hmdb_search, user_data=search_tag)
                dpg.add_button(label='X', pos=(dpg.get_viewport_width() - 40, 50), parent=tab_tag, callback=close,
                               user_data=tab_tag)

def hmdb_search(sender, app_data, user_data):
    compounds = find_xml(dpg.get_value(user_data)) # list of lists
    compound_tags = [] # list of compound tag names, strings
    windowCount.hmdb_popup_count += 1
    tab_tag = str(f'{windowCount.hmdb_popup_count}. HMDB Search Results: {user_data}')
    with dpg.window(tag=tab_tag, label=tab_tag, width=600, height=300):
        for compound in compounds:
            if compound != 'Invalid Accession':
                top_text_add_compound = str(f'top text add compound {windowCount.hmdb_popup_count}')
                dpg.add_text('Select the compound you want to add!', parent=tab_tag, tag=top_text_add_compound)
                compound_str = (f'{"Name":40}{compound.name}\n{"Formula":40}{compound.formula}\n{"InChiKey":40}'
                                f'{compound.inchikey}')
                compound_tag = str(f'{windowCount.pubchem_popup} {compound_str}')
                compound_tags.append(compound_tag)
                dpg.add_checkbox(tag=compound_tag, label=compound_str, parent=tab_tag)
                dpg.add_button(label='Add to Database, Continue to Next Step', parent=tab_tag,
                               callback=hmdb_popup_add_info, user_data=[compounds, tab_tag, compound_tags,
                                                                        top_text_add_compound])
            else:
                dpg.add_text('Invalid Accession. Try to search again')
                dpg.add_button(label='Close', callback=close, user_data=tab_tag)

def hmdb_popup_add_info(sender, app_data, user_data):
    compounds = user_data[0]
    tab_tag = user_data[1]
    compound_tags = user_data[2]
    compound_to_use = []
    dpg.delete_item(user_data[3])
    dpg.delete_item(sender)

    for x in range(len(compound_tags)):
        if dpg.get_value(compound_tags[x]) == True:
            compound_to_use.append(compounds[x])
            dpg.delete_item(compound_tags[x])
        else:
            dpg.delete_item(compound_tags[x])
    chosen_compound = compound_to_use[0]

    # TODO Ask about the type of SMILES the HMDB returns
    this_compound_information = ['Name', 'MolecularFormula', 'MolecularWeight', 'MonoisotopicMass', 'IUPACName', 'CAS',
                                 'CanonicalSMILES', 'IsomericSMILES', 'InChI', 'InChIKey']

    xl_data = gc(get_filepath())
    names = xl_data[0]
    name_tags = []
    label_tags = []
    labels = []
    for x in range(2, 47):
        if names[x] is not None and names[x] != 'STANDARD PROPERTIES <<--' and names[x] != \
                'PARENT COMPOUNT PROPERTIES -->>':
            if names[x] not in this_compound_information:
                name_tag = str(f'{windowCount.pubchem_popup}:{x}. {names[x]}')
                label = str(f'{x}. {names[x]}')
                labels.append(label)
                label_tag = str(f'{windowCount}_{label}')
                label_tags.append(label_tag)
                dpg.add_text(label,tag=label_tag, parent=tab_tag)
                dpg.add_input_text(tag=name_tag, parent=tab_tag)
                name_tags.append(name_tag)
    button_tag = str(f'{windowCount.pubchem_popup}_add_pubchem_chemical')
    dpg.add_button(label='Add Chemical to Database!', tag=button_tag,
                   parent=tab_tag, callback=add_with_hmdb,
                   user_data=[chosen_compound, name_tags, tab_tag, labels])




def add_with_hmdb(sender, app_data, user_data):
    chosen_compound = user_data[0]
    name_tags = user_data[1]
    tab_tag = user_data[2]
    labels = user_data[3]

    wb = load_workbook(get_filepath())
    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    new_row = max_row + 1

    names = []
    for x in range(1, max_col):
        names.append(sheet.cell(1, x).value)

    # creating a new accession
    prev_accession = sheet.cell(max_row, 1).value
    zeros = 6 - len(str(int(prev_accession[5::]) + 1))
    sheet.cell(new_row, 1).value = str(prev_accession[0:5]) + str('0' * zeros) + str(int(prev_accession[5::]) + 1)

    sheet.cell(new_row, 2).value = chosen_compound.name
    sheet.cell(new_row, 13).value = chosen_compound.cas
    sheet.cell(new_row, 15).value = chosen_compound.formula
    sheet.cell(new_row, 16).value = float(chosen_compound.avg_mol_weight)
    sheet.cell(new_row, 21).value = chosen_compound.formula
    sheet.cell(new_row, 22).value = float(chosen_compound.avg_mol_weight)
    # TODO ask about what type of SMILES comes from HMDB
    sheet.cell(new_row, 23).value = chosen_compound.smiles
    sheet.cell(new_row, 24).value = chosen_compound.smiles
    sheet.cell(new_row, 25).value = chosen_compound.inchi
    sheet.cell(new_row, 26).value = chosen_compound.inchikey
    sheet.cell(new_row, 27).value = chosen_compound.iupac_name
    sheet.cell(new_row, 29).value = float(chosen_compound.monoisotopic_mass)

    # 0 is for strings, 1 is for ints, 2, is for floats
    name_types = {'Single/Mixture': 0, 'Solid/Liquid/Solution': 0, 'Location': 0, 'in stock': 0, 'Cat_number': 0,
                  'Date': 0, 'vendor suggested storage': 0, 'Expiration': 0, 'cid': 1, 'TPSA': 1, 'Complexity': 1,
                  'Charge' : 1, 'HBondDonorCount': 1, 'HBondAcceptorCount': 1, 'RotatableBondCount': 1,
                  'HeavyAtomCount': 1, 'IsotopeAtomCount': 1, 'AtomStereoCount': 1, 'DefinedAtomStereoCount': 1,
                  'UndefinedAtomStereoCount': 1, 'BondStereoCount': 1, 'DefinedBondStereoCount': 1,
                  'UndefinedBondStereoCount': 1, 'CovalentUnitCount':1, 'ConformerCount3D': 1, 'Fingerprint2D': 0,
                  'XLogP': 2}

    for x in range(1, max_col):
        name = names[x - 1]
        if name is not None:
            for y in range(len(labels)):
                label = labels[y].split('.')
                cut_label = label[1].replace(' ', '')
                if cut_label.lower() in name.lower():
                    if sheet.cell(new_row, x).value is None:
                        this_type = name_types.get(name)
                        if this_type == 0:
                            sheet.cell(new_row, x).value = dpg.get_value(name_tags[y])
                        elif this_type == 1:
                            try:
                                sheet.cell(new_row, x).value = int(dpg.get_value(name_tags[y]))
                            except:
                                sheet.cell(new_row, x).value = dpg.get_value(name_tags[y])
                        elif this_type == 2:
                            try:
                                sheet.cell(new_row, x).value = float(dpg.get_value(name_tags[y]))
                            except:
                                sheet.cell(new_row, x).value = dpg.get_value(name_tags[y])

    wb.save(get_filepath())

    close_popup(tab_tag=tab_tag, message='Chemical has been added from HMDB!')

class fileWindow:
    def file_window(self):
        windowCount.file_window_count += 1
        tab_tag = str(f'file_window_{windowCount.file_window_count}')
        with dpg.tab(label='File', tag=tab_tag, parent='tabs'):
            tags = {}
            dpg.add_button(label='X', pos=(dpg.get_viewport_width() - 40, 50), parent=tab_tag, callback=close,
                           user_data=tab_tag)
            dpg.add_text('Edit Main Database Filepath', parent=tab_tag)
            main_tag = str(f'MAIN_DATABASE_{windowCount.file_window_count}')
            dpg.add_input_text(default_value=get_filepath(), tag=main_tag, parent=tab_tag)
            tags['MAIN_DATABASE'] = main_tag
            dpg.add_text('\n', parent=tab_tag)
            personal_paths_info = get_filepath(main_path=False)
            for path_name, path in personal_paths_info.items():
                if path_name != 'MAIN_DATABASE':
                    path_tag = str(f'{path_name}_{windowCount.file_window_count}')
                    tags[path_name] = path_tag
                    dpg.add_text(path_name, parent=tab_tag)
                    dpg.add_input_text(tag=path_tag, default_value=path, parent=tab_tag)
            dpg.add_button(label='Update Filepaths', parent=tab_tag, callback=edit_filepaths, user_data=[tags, tab_tag])
            dpg.add_button(label='Add Personal Path', parent=tab_tag, callback=add_personal_path_popup)

class helpWindow:
    def help_window(self):
        windowCount.help_count += 1
        tab_tag = str(f'Help {windowCount.help_count}')
        with dpg.tab(label='Help', tag=tab_tag, parent='tabs'):
            dpg.add_text(read_help(get_help_path()), parent=tab_tag)
            dpg.add_button(label='X', pos=(dpg.get_viewport_width() - 40, 50), parent=tab_tag, callback=close,
                           user_data=tab_tag)


# ----------------------------------------------------------------------------------------------------------------------
# Miscellaneous Features

def get_filepath(main_path=True):
    with open(resource_path(), 'r') as a:
        b = a.readlines()
        my_paths = {}
        for x in range(len(b)):
            if b[x] != '\n':
                path_name = (((b[x].split('='))[0]).replace(' ', '').replace('\n', ''))
                path = (((b[x].split('='))[1]).replace(' ', '').replace('\n', ''))
                my_paths[path_name] = path
        if main_path == True:
            return my_paths.get('MAIN_DATABASE')
        else:
            return my_paths

def resource_path():
    relative_path = 'filepath.txt'
    basedir = sys.executable
    last_dir = basedir.rfind("\\")
    basedir = basedir[:last_dir]
    return str(f'{basedir}\\{relative_path}')

def get_help_path():
    relative_path = 'help.txt'
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

def read_help(help_path):
    help_str = ''
    with open(help_path) as a:
        b = a.readlines()
        for line in b:
            new_line = line.replace('\n', '')
            help_str += str(f'{new_line}\n')
    return help_str


def edit_filepaths(sender, app_data, user_data):
    path_info = user_data[0]
    tab_tag = user_data[1]
    with open(resource_path(), 'w') as a:
        out_str = ''
        for path_name, path in path_info.items():
            out_str += str(f'{path_name}={dpg.get_value(path)}\n')
        a.write(out_str)
    close_popup(tab_tag=tab_tag, message='Filepaths have been updated!')

def add_personal_path_popup():
    windowCount.personal_path_popup += 1
    tab_tag = str(f'add_personal_path_{windowCount.personal_path_popup}')
    with dpg.window(label='New Database Path', tag=tab_tag):
        name_tag = str(f'{tab_tag}_name_tag')
        filepath_tag = str(f'{tab_tag}_filepath_tag')
        dpg.add_text('Enter New Database Name', parent=tab_tag)
        dpg.add_input_text(tag=name_tag, parent=tab_tag)
        dpg.add_text('Enter filepath to database with filename included.\nIf the file doesn\'t exist, enter the name '
                     'and path of the file you want to create:')
        dpg.add_input_text(tag=filepath_tag, parent=tab_tag)
        dpg.add_button(label='Add New Database!', parent=tab_tag, callback=add_path, user_data=(name_tag, filepath_tag,
                                                                                                tab_tag))

def add_path(sender, app_data, user_data):
    name = dpg.get_value(user_data[0])
    filepath = dpg.get_value(user_data[1])
    filepath_to_add = ''

    if filepath.endswith('.xlsx'):
        filepath_to_add += filepath
    else:
        filepath_to_add += str(f'{filepath}.xlsx')

    test_filepath = gc(filepath_to_add)
    if test_filepath == 'Invalid filepath':
        workbook = xlsxwriter.Workbook(filepath_to_add)
        workbook.add_worksheet()
        workbook.close()
    else:
        pass

    try:
        with open(resource_path(), 'a') as a:
            new_path_str = str(f'\n{name}={filepath_to_add}')
            a.write(new_path_str)
            a.close()
            close_popup(message='New database filepath has been added')
    except:
        close_popup(message='No filepath.txt file available. Look at setup information to create a filepath.txt file '
                            'within the same directory as the app.\nAfter this setup, you will be able to add new files'
                            ' with ease and can store their path\'s')

def add_path_no_button(name, filepath):
    filepath_to_add = ''

    if filepath.endswith('.xlsx'):
        filepath_to_add += filepath
    else:
        filepath_to_add += str(f'{filepath}.xlsx')

    test_filepath = gc(filepath_to_add)
    if test_filepath == 'Invalid filepath':
        workbook = xlsxwriter.Workbook(filepath_to_add)
        workbook.add_worksheet()
        workbook.close()
    else:
        pass

    try:
        with open(resource_path(), 'a') as a:
            new_path_str = str(f'{name}={filepath_to_add}')
            a.write(new_path_str)
            a.close()
            close_popup(message='New database filepath has been added')
    except:
        close_popup(message='No filepath.txt file available. Look at setup information to create a filepath.txt file '
                            'within the same directory as the app.\nAfter this setup, you will be able to add new files'
                            ' with ease and can store their path\'s')

def close(sender, app_data, user_data):
    dpg.delete_item(user_data)


def invalid_filepath_popup():
    windowCount.invalid_filepath_popup_count += 1
    with dpg.window(label='Invalid Filepath or No Filepath', tag=f'Invalid Filepath or No Filepath '
                                                                 f'{windowCount.invalid_filepath_popup_count}'):
        dpg.add_text('Invalid Filepath or No Filepath\nEnter a new filepath on the Home page')
        dpg.add_button(label='Close', tag=str(f'Button Invalid Filepath or No Filepath '
                                              f'{windowCount.invalid_filepath_popup_count}'), callback=close,
                       user_data=f'Invalid Filepath or No Filepath {windowCount.invalid_filepath_popup_count}')

def close_popup(message, tab_tag=None):
    if tab_tag != None:
        dpg.delete_item(tab_tag, children_only=True)
        dpg.add_text(message, parent=tab_tag)
        dpg.add_button(label='Close', parent=tab_tag, callback=close, user_data=tab_tag)
    else:
        windowCount.close_popup_count += 1
        close_tag = str(f'close_popup_{windowCount.close_popup_count}')
        with dpg.window(label='Close Window', tag=close_tag):
            dpg.add_text(message, parent=close_tag)
            dpg.add_button(label='Close', parent=close_tag, callback=close, user_data=close_tag)


def setup_filepath():
    try:
        with open(resource_path(), 'x') as a:
            a.write('MAIN_DATABASE=Enter_Your_Filepath!')
    except:
        pass





# ---------------------------------------------------------------------------------------------------------------------
# Home Screen

def main():
    dpg.create_context()
    dpg.create_viewport(title='PMF Chemical Database', width=800, height=600)
    setup_filepath()
    with dpg.window(label='PMF Chemical Database', width=600, height=300, tag='Primary Window'):
        with dpg.menu_bar():
            dpg.add_menu_item(label='Help', tag='Help', callback=helpWindow.help_window)
            dpg.add_menu_item(label='File', tag='File', callback=fileWindow.file_window)
            dpg.add_menu_item(tag='search', label='Search Database', callback=searchWindow.search_window)
            with dpg.menu(label='Edit Database'):
                with dpg.menu(label='Add to Database'):
                    dpg.add_menu_item(label='Add Manually', callback=manualTabAdd.manual_window)
                    dpg.add_menu_item(label='Add from PubChem', callback=pubChemAdd.pubchem_window)
                    dpg.add_menu_item(label='Add from HMDB', callback=hmdbAdd.hmdb_window)
            with dpg.tab_bar(tag='tabs', parent='Primary Window'):
                with dpg.tab(label='Home', tag='Home'):
                    dpg.add_text('Welcome to the Database!', parent='Home')
    dpg.setup_dearpygui()
    dpg.show_viewport()
    dpg.set_primary_window("Primary Window", True)
    dpg.start_dearpygui()
    dpg.destroy_context()


if __name__ == '__main__':
    main()


