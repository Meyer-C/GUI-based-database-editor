from openpyxl import load_workbook

def get_compounds(filepath):

    try:

        wb = load_workbook(filepath)
        sheet = wb.active
        # for some reason the rows and columns are switched :\
        all_row = []
        for value in sheet.iter_rows(values_only = True):
            if list(value) != None:
                all_row.append(list(value))

        return list(all_row)

    except:

        return 'Invalid filepath'

print(len(get_compounds(r'C:\Users\Camden\Downloads\library_subset.xlsx')[0]))


    

